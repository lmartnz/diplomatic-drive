import streamlit as st
import pandas as pd
from datetime import datetime
import pytz
from streamlit_gsheets import GSheetsConnection
import io
from openpyxl import load_workbook

# --- CONFIGURACIÓN DE PÁGINA ---
st.set_page_config(page_title="Diplomatic Drive", page_icon="🚗", layout="wide")

# --- CONEXIÓN A GOOGLE SHEETS ---
conn = st.connection("gsheets", type=GSheetsConnection)

# --- FUNCIONES DE BASE DE DATOS ---
def cargar_datos():
    try:
        return conn.read(worksheet="Hoja 1", ttl=0)
    except:
        return pd.DataFrame()

def cargar_configuracion():
    try:
        df_config = conn.read(worksheet="config", ttl=0)
        return dict(zip(df_config.iloc[:, 0], df_config.iloc[:, 1]))
    except:
        return {}

def guardar_viaje(datos):
    try:
        df_actual = cargar_datos()
        nuevo_df = pd.DataFrame([datos])
        df_actualizado = pd.concat([df_actual, nuevo_df], ignore_index=True)
        conn.update(worksheet="Hoja 1", data=df_actualizado)
        return True
    except Exception as e:
        st.error(f"Error guardando en la nube: {e}")
        return False

# --- FUNCIONES DE TIEMPO (Washington DC) ---
def obtener_hora_actual():
    zona_dc = pytz.timezone('America/New_York')
    return datetime.now(zona_dc).strftime("%H:%M")

def obtener_timestamp_dc():
    zona_dc = pytz.timezone('America/New_York')
    return str(datetime.now(zona_dc))

# --- CALLBACKS ---
def set_hora_salida():
    st.session_state.hora_salida = obtener_hora_actual()

def set_hora_llegada():
    st.session_state.hora_llegada = obtener_hora_actual()

# ==========================================
# 🔽 MENÚ LATERAL 🔽
# ==========================================

with st.sidebar:
    st.image("https://upload.wikimedia.org/wikipedia/commons/thumb/f/f2/Flag_of_Costa_Rica.svg/320px-Flag_of_Costa_Rica.svg.png", width=100)
    st.title("Diplomatic Drive")
    st.subheader("Misión Permanente OEA")
    
    menu = st.radio(
        "Menú Principal",
        ["🏠 Inicio", "📅 Agenda", "🚗 Bitácora Oficial", "📄 Reportes Cancillería", "⚙️ Mantenimiento"],
        index=0
    )
    
    st.markdown("---")
    st.caption("🟢 Sistema En Línea | V 2.1 Mapeo Corregido")

# ==========================================
# 🔽 LÓGICA DE PÁGINAS 🔽
# ==========================================

# 1. INICIO
if menu == "🏠 Inicio":
    st.title("Bienvenido, Luis")
    st.markdown("### 🛡️ Panel de Control Oficial")
    st.info("Sistema listo. Seleccione 'Bitácora Oficial' para registrar movimientos.")
    
    df = cargar_datos()
    if not df.empty:
        st.metric("Total Viajes Registrados", len(df))

# 2. AGENDA
elif menu == "📅 Agenda":
    st.title("📅 Agenda de Movimientos")
    st.write("🚧 Módulo de Integración con Outlook (Fase 2)")

# 3. BITÁCORA
elif menu == "🚗 Bitácora Oficial":
    st.title("📒 Registro de Movimientos")
    st.markdown("*Formulario conectado a Base de Datos Segura*")
    
    # Memoria Inteligente
    df = cargar_datos()
    def_lugar_salida = ""
    def_odo_inicial = 0
    
    if not df.empty:
        ultimo_viaje = df.iloc[-1]
        def_lugar_salida = ultimo_viaje.get("lugar_llegada", "")
        try:
            def_odo_inicial = int(ultimo_viaje.get("odo_final", 0))
        except:
            def_odo_inicial = 0
    
    with st.form("entry_form", clear_on_submit=False):
        col1, col2 = st.columns(2)
        
        with col1:
            fecha = st.date_input("Fecha del Viaje", datetime.now())
            st.write("---")
            
            col_b1, col_i1 = st.columns([1,2])
            with col_b1:
                st.form_submit_button("🕒 Salida", on_click=set_hora_salida, type="secondary")
            with col_i1:
                if 'hora_salida' not in st.session_state: st.session_state.hora_salida = ""
                hora_sal = st.text_input("Hora Salida", key='hora_salida')
            
            lugar_sal = st.text_input("📍 Lugar Salida", value=def_lugar_salida)
            odo_ini = st.number_input("🔢 Odómetro Inicial", min_value=0, value=def_odo_inicial)

        with col2:
            st.write("")
            st.write("---")
            
            col_b2, col_i2 = st.columns([1,2])
            with col_b2:
                st.form_submit_button("🏁 Llegada", on_click=set_hora_llegada, type="secondary")
            with col_i2:
                if 'hora_llegada' not in st.session_state: st.session_state.hora_llegada = ""
                hora_lle = st.text_input("Hora Llegada", key='hora_llegada')

            lugar_lle = st.text_input("📍 Lugar Llegada")
            odo_fin = st.number_input("🔢 Odómetro Final", min_value=0)

        st.write("---")
        asunto = st.text_area("📝 Asunto / Misión (Detalle para reporte)")
        costo = st.number_input("💵 Gastos ($)", min_value=0.0, format="%.2f")

        submitted = st.form_submit_button("💾 GUARDAR EN NUBE", type="primary")
        
        if submitted:
            if not asunto:
                st.error("⚠️ Falta el Asunto.")
            elif odo_fin < odo_ini and odo_fin != 0:
                 st.error(f"⚠️ Error: El odómetro final no puede ser menor al inicial.")
            else:
                # FORMATO LATINO PARA GUARDADO
                nuevo_registro = {
                    "fecha": fecha.strftime("%d/%m/%Y"), 
                    "hora_salida": str(hora_sal),
                    "lugar_salida": lugar_sal,
                    "odo_inicial": int(odo_ini),
                    "hora_llegada": str(hora_lle),
                    "lugar_llegada": lugar_lle,
                    "odo_final": int(odo_fin),
                    "costo": float(costo),
                    "asunto": asunto,
                    "timestamp_registro": obtener_timestamp_dc()
                }
                
                with st.spinner("Encriptando y guardando..."):
                    if guardar_viaje(nuevo_registro):
                        st.success("✅ ¡Viaje registrado exitosamente!")
                        st.balloons()
                        st.rerun()

# 4. REPORTES (CORREGIDO: Mapeo exacto de columnas)
elif menu == "📄 Reportes Cancillería":
    st.title("🖨️ Centro de Reportes")
    st.markdown("### Generación de Informe en Formato Oficial")
    st.info("🔒 Los datos se inyectarán en la plantilla `plantilla_oficial.xlsx`")
    
    st.write("---")
    
    col_f1, col_f2 = st.columns(2)
    with col_f1:
        inicio = st.date_input("Desde:", value=datetime.now().date().replace(day=1))
    with col_f2:
        fin = st.date_input("Hasta:", value=datetime.now().date())
        
    st.write("")
    
    if st.button("📊 GENERAR INFORME OFICIAL"):
        df = cargar_datos()
        config = cargar_configuracion()
        
        if not df.empty:
            try:
                # 1. Filtrado de Fechas
                # dayfirst=True ayuda a mezclar formatos viejos (05/01/2026) y nuevos
                df["fecha_dt"] = pd.to_datetime(df["fecha"], dayfirst=True, errors='coerce').dt.date
                df = df.dropna(subset=["fecha_dt"])
                
                mask = (df["fecha_dt"] >= inicio) & (df["fecha_dt"] <= fin)
                df_filtrado = df.loc[mask]
                
                if not df_filtrado.empty:
                    # Ordenamos por fecha para que salgan en orden
                    df_filtrado = df_filtrado.sort_values(by="fecha_dt")
                    
                    # 2. Cargar la Plantilla
                    try:
                        wb = load_workbook("plantilla_oficial.xlsx")
                        ws = wb.active
                    except FileNotFoundError:
                        st.error("❌ No se encontró 'plantilla_oficial.xlsx'. Súbela a GitHub.")
                        st.stop()

                    # 3. Llenar Encabezados (Opcional - Si usas hoja config)
                    # if config:
                    #     ws['B3'] = config.get('jefe_mision', '') # Ejemplo

                    # 4. Inyectar los Viajes (MAPEO CORREGIDO SEGÚN TU FOTO)
                    FILA_INICIO = 16  # <-- Ajustado según tu captura de pantalla
                    
                    for i, row in df_filtrado.iterrows():
                        fila_excel = FILA_INICIO + i
                        
                        # --- ASIGNACIÓN EXACTA DE COLUMNAS ---
                        # Col A (1): Fecha
                        ws.cell(row=fila_excel, column=1).value = row['fecha']
                        
                        # Col B (2): Odómetro Inicial (Antes pusimos hora aquí por error)
                        ws.cell(row=fila_excel, column=2).value = row['odo_inicial']
                        
                        # Col C (3): Lugar Salida
                        ws.cell(row=fila_excel, column=3).value = row['lugar_salida']
                        
                        # Col D (4): Hora Salida (Antes pusimos odómetro aquí)
                        ws.cell(row=fila_excel, column=4).value = row['hora_salida']
                        
                        # Col E (5): Odómetro Final
                        ws.cell(row=fila_excel, column=5).value = row['odo_final']
                        
                        # Col F (6): Lugar Llegada
                        ws.cell(row=fila_excel, column=6).value = row['lugar_llegada']
                        
                        # Col G (7): Hora Llegada
                        ws.cell(row=fila_excel, column=7).value = row['hora_llegada']
                        
                        # Col H (8): Km Recorridos (CÁLCULO AUTOMÁTICO)
                        try:
                            recorrido = int(row['odo_final']) - int(row['odo_inicial'])
                            ws.cell(row=fila_excel, column=8).value = recorrido
                        except:
                            ws.cell(row=fila_excel, column=8).value = 0
                            
                        # Col I (9): Costo Moneda Local (Opcional)
                        # ws.cell(row=fila_excel, column=9).value = "" 
                        
                        # Col J (10): Costo US$
                        ws.cell(row=fila_excel, column=10).value = row['costo']
                        
                        # Col M (13): Motivo/Justificación (La columna ancha del final)
                        ws.cell(row=fila_excel, column=13).value = row['asunto']

                    # 5. Guardar y Descargar
                    buffer = io.BytesIO()
                    wb.save(buffer)
                    valioso_excel = buffer.getvalue()
                    
                    st.success(f"✅ Informe generado con {len(df_filtrado)} registros.")
                    
                    nombre_archivo = f"Informe_Oficial_{inicio}_{fin}.xlsx"
                    st.download_button(
                        label="⬇️ DESCARGAR EXCEL OFICIAL",
                        data=valioso_excel,
                        file_name=nombre_archivo,
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        type="primary"
                    )
                else:
                    st.warning(f"⚠️ No se encontraron viajes entre {inicio} y {fin}. Verifica las fechas.")
            
            except Exception as e:
                st.error(f"Error técnico: {e}")
        else:
            st.error("Base de datos vacía.")

# 5. MANTENIMIENTO
elif menu == "⚙️ Mantenimiento":
    st.title("⚙️ Taller y Mantenimiento")
    st.write("Próximamente.")
